#!/usr/bin/env python3
"""
Create Campaign Plan Script

Generates an 8-sheet Excel workbook for campaign master plan.
Includes timelines, budgets, Gantt charts, and tracking.

Usage: python create_campaign_plan.py <workflow_state.json> <output_path>

Requires: openpyxl
"""

import json
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, Any, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


# Style definitions
STYLES = {
    'header': {
        'font': Font(bold=True, color='FFFFFF', size=12),
        'fill': PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center')
    },
    'subheader': {
        'font': Font(bold=True, size=11),
        'fill': PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    },
    'title': {
        'font': Font(bold=True, size=16, color='1F2937')
    },
    'currency': '#,##0',
    'percent': '0.0%',
    'date': 'YYYY-MM-DD'
}

PHASE_COLORS = [
    'DBEAFE',  # Blue - Phase 1
    'D1FAE5',  # Green - Phase 2
    'FEF3C7',  # Yellow - Phase 3
    'FCE7F3',  # Pink - Phase 4
    'E0E7FF',  # Indigo - Phase 5
    'FEE2E2',  # Red - Phase 6
    'CCFBF1',  # Teal - Phase 7
    'F3E8FF',  # Purple - Phase 8
]


def apply_style(cell, style_name):
    """Apply a predefined style to a cell."""
    style = STYLES.get(style_name, {})
    for attr, value in style.items():
        setattr(cell, attr, value)


def create_executive_summary(wb, data):
    """Create Executive Summary sheet."""
    ws = wb.create_sheet("Executive Summary", 0)
    
    # Title
    ws['A1'] = "CAMPAIGN MASTER PLAN"
    apply_style(ws['A1'], 'title')
    ws.merge_cells('A1:F1')
    
    campaign_name = data.get('brand_dna', {}).get('company_info', {}).get('name', 'Campaign')
    ws['A3'] = f"Campaign: {campaign_name} Marketing Campaign"
    
    # Campaign Overview Section
    ws['A5'] = "CAMPAIGN OVERVIEW"
    apply_style(ws['A5'], 'header')
    ws.merge_cells('A5:B5')
    
    overview_data = [
        ('Campaign Name:', f"{campaign_name} Marketing Campaign 2024"),
        ('Duration:', f"{data.get('campaign_parameters', {}).get('campaign_duration_months', 12)} months"),
        ('Start Date:', datetime.now().strftime('%Y-%m-%d')),
        ('End Date:', (datetime.now() + timedelta(days=365)).strftime('%Y-%m-%d')),
        ('Total Budget:', data.get('strategy_artifacts', {}).get('budget_allocation', {}).get('total_budget_range', '$100,000 - $150,000')),
        ('Primary Objective:', 'Increase brand awareness and lead generation'),
    ]
    
    for i, (label, value) in enumerate(overview_data, start=6):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Key Milestones Section
    ws['A14'] = "KEY MILESTONES"
    apply_style(ws['A14'], 'header')
    ws.merge_cells('A14:D14')
    
    ws['A15'] = "Milestone"
    ws['B15'] = "Target Date"
    ws['C15'] = "Owner"
    ws['D15'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}15'], 'subheader')
    
    milestones = [
        ('Campaign Launch', (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d'), 'Marketing Lead', 'Planned'),
        ('Content Pipeline Ready', (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'), 'Content Team', 'Planned'),
        ('Q1 Review', (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d'), 'Marketing Lead', 'Planned'),
        ('Mid-Campaign Review', (datetime.now() + timedelta(days=180)).strftime('%Y-%m-%d'), 'Leadership', 'Planned'),
        ('Campaign Completion', (datetime.now() + timedelta(days=365)).strftime('%Y-%m-%d'), 'Marketing Lead', 'Planned'),
    ]
    
    for i, (milestone, date, owner, status) in enumerate(milestones, start=16):
        ws[f'A{i}'] = milestone
        ws[f'B{i}'] = date
        ws[f'C{i}'] = owner
        ws[f'D{i}'] = status
    
    # Success Metrics Section
    ws['A23'] = "SUCCESS METRICS"
    apply_style(ws['A23'], 'header')
    ws.merge_cells('A23:D23')
    
    ws['A24'] = "KPI"
    ws['B24'] = "Target"
    ws['C24'] = "Current"
    ws['D24'] = "Status"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}24'], 'subheader')
    
    kpis = [
        ('Website Traffic', '50,000/mo', '35,000/mo', 'In Progress'),
        ('Lead Generation', '500/quarter', '320/quarter', 'In Progress'),
        ('Conversion Rate', '3.5%', '2.8%', 'In Progress'),
        ('Email Subscribers', '10,000', '7,500', 'In Progress'),
        ('Social Engagement', '5% rate', '3.2% rate', 'In Progress'),
    ]
    
    for i, (kpi, target, current, status) in enumerate(kpis, start=25):
        ws[f'A{i}'] = kpi
        ws[f'B{i}'] = target
        ws[f'C{i}'] = current
        ws[f'D{i}'] = status
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    return ws


def create_timeline_sheet(wb, data):
    """Create Campaign Timeline (Gantt) sheet."""
    ws = wb.create_sheet("Campaign Timeline")
    
    # Headers
    headers = ['Task', 'Owner', 'Start', 'End', 'Duration', 'Status']
    weeks = [f'W{i}' for i in range(1, 53)]
    all_headers = headers + weeks
    
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, 'header')
    
    # Campaign phases and tasks
    phases = [
        ('Phase 1: Foundation', 'Marketing Lead', 0, 30, 'Phase'),
        ('  Brand Guidelines', 'Brand Team', 0, 14, 'Task'),
        ('  Asset Creation', 'Design', 7, 30, 'Task'),
        ('Phase 2: Content Setup', 'Content Lead', 14, 60, 'Phase'),
        ('  Content Calendar', 'Content Team', 14, 21, 'Task'),
        ('  Blog Posts (10)', 'Writers', 21, 60, 'Task'),
        ('Phase 3: Channel Launch', 'Marketing Lead', 30, 90, 'Phase'),
        ('  SEO Optimization', 'SEO Team', 30, 60, 'Task'),
        ('  Social Media Setup', 'Social Team', 45, 75, 'Task'),
        ('  Email Templates', 'Email Team', 60, 90, 'Task'),
        ('Phase 4: Campaign Execution', 'Marketing Lead', 60, 180, 'Phase'),
        ('  Content Publishing', 'Content Team', 60, 180, 'Task'),
        ('  Paid Ads Launch', 'Ads Team', 75, 120, 'Task'),
        ('  Email Campaigns', 'Email Team', 90, 180, 'Task'),
        ('Phase 5: Optimization', 'Marketing Lead', 120, 270, 'Phase'),
        ('  Performance Review', 'Analytics', 120, 135, 'Task'),
        ('  A/B Testing', 'All Teams', 135, 180, 'Task'),
        ('  Content Refresh', 'Content Team', 180, 270, 'Task'),
        ('Phase 6: Scale', 'Marketing Lead', 270, 365, 'Phase'),
        ('  Budget Increase', 'Finance', 270, 285, 'Task'),
        ('  Channel Expansion', 'Marketing', 285, 365, 'Task'),
    ]
    
    start_date = datetime.now()
    
    for row, (task, owner, start_day, end_day, task_type) in enumerate(phases, 2):
        ws.cell(row=row, column=1, value=task)
        ws.cell(row=row, column=2, value=owner)
        ws.cell(row=row, column=3, value=(start_date + timedelta(days=start_day)).strftime('%Y-%m-%d'))
        ws.cell(row=row, column=4, value=(start_date + timedelta(days=end_day)).strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=end_day - start_day)
        ws.cell(row=row, column=6, value='Planned')
        
        # Gantt bars
        start_week = start_day // 7 + 1
        end_week = end_day // 7 + 1
        
        fill_color = PHASE_COLORS[row % len(PHASE_COLORS)] if task_type == 'Phase' else 'BFDBFE'
        
        for week in range(start_week, min(end_week + 1, 53)):
            cell = ws.cell(row=row, column=6 + week)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.value = '█'
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    
    for i in range(7, 59):
        ws.column_dimensions[get_column_letter(i)].width = 4
    
    return ws


def create_budget_sheet(wb, data):
    """Create Budget Allocation sheet."""
    ws = wb.create_sheet("Budget Allocation")
    
    # Headers
    headers = ['Category', 'Subcategory', 'Monthly', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, 'header')
    
    # Budget data
    budget_items = [
        ('Content Marketing', 'Blog Posts', 2500, 7500, 7500, 7500, 7500),
        ('Content Marketing', 'Whitepapers', 1500, 4500, 4500, 4500, 4500),
        ('Content Marketing', 'Video Production', 3000, 9000, 9000, 9000, 9000),
        ('Paid Advertising', 'Google Ads', 4000, 12000, 12000, 12000, 12000),
        ('Paid Advertising', 'LinkedIn Ads', 2500, 7500, 7500, 7500, 7500),
        ('Paid Advertising', 'Social Ads', 1500, 4500, 4500, 4500, 4500),
        ('SEO', 'Tools & Software', 500, 1500, 1500, 1500, 1500),
        ('SEO', 'Link Building', 1000, 3000, 3000, 3000, 3000),
        ('Email Marketing', 'Platform', 300, 900, 900, 900, 900),
        ('Email Marketing', 'List Building', 700, 2100, 2100, 2100, 2100),
        ('Events', 'Webinars', 1000, 3000, 3000, 3000, 3000),
        ('Events', 'Conferences', 2000, 6000, 6000, 6000, 6000),
        ('Tools & Software', 'Analytics', 400, 1200, 1200, 1200, 1200),
        ('Tools & Software', 'Automation', 600, 1800, 1800, 1800, 1800),
        ('Contingency', 'Reserve (10%)', 2000, 6000, 6000, 6000, 6000),
    ]
    
    row = 2
    for category, subcategory, monthly, q1, q2, q3, q4 in budget_items:
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=subcategory)
        ws.cell(row=row, column=3, value=monthly).number_format = STYLES['currency']
        ws.cell(row=row, column=4, value=q1).number_format = STYLES['currency']
        ws.cell(row=row, column=5, value=q2).number_format = STYLES['currency']
        ws.cell(row=row, column=6, value=q3).number_format = STYLES['currency']
        ws.cell(row=row, column=7, value=q4).number_format = STYLES['currency']
        ws.cell(row=row, column=8, value=q1+q2+q3+q4).number_format = STYLES['currency']
        row += 1
    
    # Totals row
    ws.cell(row=row, column=1, value='TOTAL')
    apply_style(ws.cell(row=row, column=1), 'subheader')
    for col in range(3, 9):
        cell = ws.cell(row=row, column=col)
        cell.value = f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{row-1})'
        cell.number_format = STYLES['currency']
        apply_style(cell, 'subheader')
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    
    return ws


def create_channel_schedule(wb, data):
    """Create Channel Schedule sheet."""
    ws = wb.create_sheet("Channel Schedule")
    
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    headers = ['Channel', 'Activity', 'Frequency'] + months
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, 'header')
    
    schedule = [
        ('SEO', 'Blog Posts', '4/month', [1]*12),
        ('SEO', 'Technical SEO', 'Monthly', [1]*12),
        ('SEO', 'Link Building', 'Ongoing', [1]*12),
        ('Email', 'Newsletter', '2/month', [1]*12),
        ('Email', 'Drip Campaign', 'Ongoing', [1]*12),
        ('Email', 'Promotional', 'Monthly', [1]*12),
        ('Social', 'LinkedIn Posts', '3/week', [1]*12),
        ('Social', 'Twitter/X', 'Daily', [1]*12),
        ('Social', 'Instagram', '2/week', [1]*12),
        ('Paid', 'Google Ads', 'Daily', [1]*12),
        ('Paid', 'LinkedIn Ads', 'Weekdays', [1]*12),
        ('Paid', 'Retargeting', 'Ongoing', [1]*12),
        ('Events', 'Webinars', 'Monthly', [1]*12),
        ('Events', 'Conferences', 'Quarterly', [1,0,0,1,0,0,1,0,0,1,0,0]),
    ]
    
    for row, (channel, activity, freq, months_active) in enumerate(schedule, 2):
        ws.cell(row=row, column=1, value=channel)
        ws.cell(row=row, column=2, value=activity)
        ws.cell(row=row, column=3, value=freq)
        
        for col, active in enumerate(months_active, 4):
            cell = ws.cell(row=row, column=col)
            if active:
                cell.value = '✓'
                cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    for col in range(4, 16):
        ws.column_dimensions[get_column_letter(col)].width = 6
    
    return ws


def create_content_calendar(wb, data):
    """Create Content Calendar sheet."""
    ws = wb.create_sheet("Content Calendar")
    
    headers = ['Date', 'Content Type', 'Title', 'Topic', 'Channel', 'Author', 'Status', 'Due Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, 'header')
    
    # Sample content calendar
    start_date = datetime.now()
    content_types = ['Blog Post', 'Whitepaper', 'Case Study', 'Video', 'Infographic', 'Newsletter']
    topics = ['Industry Trends', 'Product Update', 'How-To Guide', 'Customer Story', 'Thought Leadership']
    channels = ['SEO', 'Email', 'Social', 'Gated Content', 'YouTube']
    statuses = ['Planned', 'Outline', 'Draft', 'Review', 'Final']
    
    for i in range(30):
        row = i + 2
        content_date = start_date + timedelta(days=i*7)
        
        ws.cell(row=row, column=1, value=content_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=content_types[i % len(content_types)])
        ws.cell(row=row, column=3, value=f'Content Piece {i+1}')
        ws.cell(row=row, column=4, value=topics[i % len(topics)])
        ws.cell(row=row, column=5, value=channels[i % len(channels)])
        ws.cell(row=row, column=6, value='Content Team')
        ws.cell(row=row, column=7, value=statuses[0])
        ws.cell(row=row, column=8, value=(content_date - timedelta(days=7)).strftime('%Y-%m-%d'))
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    
    return ws


def create_kpi_tracking(wb, data):
    """Create KPI Tracking sheet."""
    ws = wb.create_sheet("KPI Tracking")
    
    months = ['Target', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'YTD']
    headers = ['KPI'] + months + ['Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, 'header')
    
    kpis = [
        ('Website Traffic', 50000, [42000, 44000, 46000, 48000, 50000, 52000, 54000, 56000, 58000, 60000, 62000, 64000]),
        ('Lead Generation', 500, [380, 400, 420, 440, 460, 480, 500, 520, 540, 560, 580, 600]),
        ('Conversion Rate', 0.035, [0.028, 0.029, 0.030, 0.031, 0.032, 0.033, 0.034, 0.035, 0.036, 0.037, 0.038, 0.039]),
        ('Email Subscribers', 10000, [7500, 7800, 8100, 8400, 8700, 9000, 9300, 9600, 9900, 10200, 10500, 10800]),
        ('Social Followers', 5000, [3500, 3700, 3900, 4100, 4300, 4500, 4700, 4900, 5100, 5300, 5500, 5700]),
        ('Engagement Rate', 0.05, [0.032, 0.034, 0.036, 0.038, 0.040, 0.042, 0.044, 0.046, 0.048, 0.050, 0.052, 0.054]),
    ]
    
    for row, (kpi, target, monthly_values) in enumerate(kpis, 2):
        ws.cell(row=row, column=1, value=kpi)
        ws.cell(row=row, column=2, value=target)
        
        if 'Rate' in kpi:
            ws.cell(row=row, column=2).number_format = STYLES['percent']
        
        ytd = 0
        for col, value in enumerate(monthly_values, 3):
            cell = ws.cell(row=row, column=col, value=value)
            if 'Rate' in kpi:
                cell.number_format = STYLES['percent']
            ytd += value if 'Rate' not in kpi else value / 12
        
        ws.cell(row=row, column=15, value=ytd if 'Rate' not in kpi else ytd)
        if 'Rate' in kpi:
            ws.cell(row=row, column=15).number_format = STYLES['percent']
        
        # Status
        status = '🟢 On Track' if monthly_values[0] >= target * 0.8 else '🟡 At Risk'
        ws.cell(row=row, column=16, value=status)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    for col in range(2, 17):
        ws.column_dimensions[get_column_letter(col)].width = 10
    
    return ws


def create_resource_allocation(wb, data):
    """Create Resource Allocation sheet."""
    ws = wb.create_sheet("Resource Allocation")
    
    headers = ['Resource', 'Role', 'Allocation', 'Tasks', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, 'header')
    
    resources = [
        ('Marketing Lead', 'Campaign Owner', '100%', 'Strategy, Oversight', [100]*12),
        ('Content Manager', 'Content Lead', '100%', 'Content Calendar', [100]*12),
        ('SEO Specialist', 'SEO', '80%', 'Optimization', [80]*12),
        ('Social Media Manager', 'Social', '100%', 'Social Channels', [100]*12),
        ('Email Marketing Specialist', 'Email', '80%', 'Email Campaigns', [80]*12),
        ('Graphic Designer', 'Design', '60%', 'Visual Assets', [60]*12),
        ('Copywriter', 'Content', '100%', 'Blog, Copy', [100]*12),
        ('Analytics Specialist', 'Analytics', '50%', 'Reporting', [50]*12),
    ]
    
    for row, (name, role, alloc, tasks, monthly) in enumerate(resources, 2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=role)
        ws.cell(row=row, column=3, value=alloc)
        ws.cell(row=row, column=4, value=tasks)
        
        for col, value in enumerate(monthly, 5):
            cell = ws.cell(row=row, column=col, value=f'{value}%')
            if value >= 80:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            elif value >= 50:
                cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    for col in range(5, 17):
        ws.column_dimensions[get_column_letter(col)].width = 6
    
    return ws


def create_risk_register(wb, data):
    """Create Risk Register sheet."""
    ws = wb.create_sheet("Risk Register")
    
    headers = ['Risk ID', 'Risk Description', 'Probability', 'Impact', 'Score', 'Mitigation', 'Owner', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_style(cell, 'header')
    
    risks = [
        ('R001', 'Budget overrun in paid advertising', 'Medium', 'High', 6, 'Monthly budget reviews', 'Marketing Lead', 'Open'),
        ('R002', 'Resource shortage during peak periods', 'Low', 'Medium', 2, 'Contractor backup list', 'Marketing Lead', 'Mitigated'),
        ('R003', 'Content quality inconsistency', 'Medium', 'Medium', 4, 'Style guide & review process', 'Content Manager', 'Open'),
        ('R004', 'Platform algorithm changes', 'High', 'Medium', 6, 'Diversified channel strategy', 'SEO Specialist', 'Monitoring'),
        ('R005', 'Low conversion rates', 'Medium', 'High', 6, 'A/B testing program', 'Analytics', 'Open'),
        ('R006', 'Competitor campaign response', 'Medium', 'Low', 3, 'Monitoring & differentiation', 'Marketing Lead', 'Monitoring'),
    ]
    
    for row, (rid, desc, prob, impact, score, mitigation, owner, status) in enumerate(risks, 2):
        ws.cell(row=row, column=1, value=rid)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=prob)
        ws.cell(row=row, column=4, value=impact)
        ws.cell(row=row, column=5, value=score)
        ws.cell(row=row, column=6, value=mitigation)
        ws.cell(row=row, column=7, value=owner)
        ws.cell(row=row, column=8, value=status)
        
        # Color code risk score
        score_cell = ws.cell(row=row, column=5)
        if score >= 6:
            score_cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        elif score >= 4:
            score_cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        else:
            score_cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    
    return ws


def create_campaign_plan(workflow_state_path: str, output_path: str) -> Dict[str, Any]:
    """Main function to create the campaign plan Excel file."""
    
    # Load workflow state
    with open(workflow_state_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all sheets
    create_executive_summary(wb, data)
    create_timeline_sheet(wb, data)
    create_budget_sheet(wb, data)
    create_channel_schedule(wb, data)
    create_content_calendar(wb, data)
    create_kpi_tracking(wb, data)
    create_resource_allocation(wb, data)
    create_risk_register(wb, data)
    
    # Save workbook
    wb.save(output_path)
    
    return {
        'status': 'success',
        'output_path': output_path,
        'sheets_created': 8,
        'created_at': datetime.now().isoformat()
    }


def main():
    """CLI entry point."""
    if len(sys.argv) < 3:
        print("Usage: python create_campaign_plan.py <workflow_state.json> <output_path>")
        print("\nExample: python create_campaign_plan.py workflow_state.json Campaign_Master_Plan.xlsx")
        sys.exit(1)
    
    workflow_state_path = sys.argv[1]
    output_path = sys.argv[2]
    
    result = create_campaign_plan(workflow_state_path, output_path)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()